# import pandas as pd
# from openpyxl import Workbook
# from openpyxl.drawing.image import Image
# from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
# from openpyxl.utils import get_column_letter
# import tkinter as tk
# from tkinter import filedialog, messagebox

# # Function to generate exam desk cards
# def generate_exam_desk_cards():
#     file_path = "test_seat_plan1.xlsx"  # Default file name
#     # Read the data from the Excel file
#     try:
#         df = pd.read_excel(file_path)
#     except Exception as e:
#         messagebox.showerror("Error", f"Error reading file: {e}")
#         return

#     # Group students by exam venue and session
#     grouped_data = df.groupby(['Venue', 'Session', 'Board'])

#     # Define colors for each unique paper code
#     color_dict = {}
#     colors = [PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
#               PatternFill(start_color='FFEB9B', end_color='FFEB9B', fill_type='solid'),
#               PatternFill(start_color='C9FFC3', end_color='C9FFC3', fill_type='solid'),
#               PatternFill(start_color='9BEBFF', end_color='9BEBFF', fill_type='solid'),
#               PatternFill(start_color='FFC3E8', end_color='FFC3E8', fill_type='solid'),
#               PatternFill(start_color='B8B8FF', end_color='B8B8FF', fill_type='solid'),
#               PatternFill(start_color='D8BFD8', end_color='D8BFD8', fill_type='solid'),
#               PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid'),
#               PatternFill(start_color='F5DEB3', end_color='F5DEB3', fill_type='solid'),
#               PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid'),
#               PatternFill(start_color='808080', end_color='808080', fill_type='solid'),
#               PatternFill(start_color='FFA07A', end_color='FFA07A', fill_type='solid'),
#               PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid'),
#               PatternFill(start_color='7B68EE', end_color='7B68EE', fill_type='solid'),
#               PatternFill(start_color='FF6347', end_color='FF6347', fill_type='solid')]

#     # Iterate over each group and generate exam desk cards
#     for (venue, session, Board), group in grouped_data:
#         # Create a file name for the exam desk cards
#         file_name = f'Seat_Planing_for_{venue}_{session}.xlsx'
        
#         # Create a workbook
#         wb = Workbook()
#         ws = wb.active
        
#         # Add logo to the worksheet with adjusted position
#         img = Image("British_Council_Logo.png")
#         img.width = 150  # Adjust the width of the image (optional)
#         img.height = 80  # Adjust the height of the image (optional)
#         ws.add_image(img, 'A1')  # Set the anchor to 'A1' to position the image within the cell
        
#         # Add headers and other details
#         ws['A5'] = f"{Board} Examinations - {group['Date'].iloc[0].strftime('%Y-%m-%d')}"
#         ws['A6'] = venue
#         ws['A7'] = f"Session - {session}"
#         ws['A8'] = f"Date: {group['Date'].iloc[0].strftime('%Y-%m-%d')}"
        
#         paper_codes = ', '.join(map(str, group['Paper code'].unique()))
#         ws['A9'] = f"Subject: {paper_codes}"

#         # Add total candidate count
#         ws.merge_cells('A10:B10')
#         ws['A10'] = "Total candidate"
#         ws['A10'].alignment = Alignment(horizontal='center', vertical='center')
#         ws['C10'] = len(group)

#         # Initialize variables for row and column numbers
#         start_row = 0
#         start_col = 1
        
#         # Iterate over candidate numbers and assign them to cells based on seat locator
#         for _, row_data in group.iterrows():
#             candidate_num = row_data['Candidate Number']
#             seat_locator = row_data['Seat Locator']
#             paper_code = row_data['Paper code']

#             if pd.notnull(seat_locator):
#                 # Convert seat locator to row and column indices
#                 col = ord(seat_locator[0].lower()) - ord('a') + start_col  # Convert column letter to index
#                 row = int(seat_locator[1:]) + start_row  # Convert row number to index
                
#                 # Write candidate number to the specified cell
#                 ws.cell(row=row, column=col).value = candidate_num
                
#                 # Add border to the cell
#                 ws.cell(row=row, column=col).border = Border(left=Side(style='thin'), 
#                                                               right=Side(style='thin'), 
#                                                               top=Side(style='thin'), 
#                                                               bottom=Side(style='thin'))
 
#                 # Assigning color to the cell based on the paper code
#                 if paper_code not in color_dict:
#                     color_dict[paper_code] = colors.pop(0)
                
#                 # Applying fill color to the cell
#                 ws.cell(row=row, column=col).fill = color_dict[paper_code]

#         # Add total candidate count for each paper code
#         for i, paper_code in enumerate(group['Paper code'].unique(), start=11):
#             count = len(group[group['Paper code'] == paper_code])
#             if count > 0:
#                 # Write total count for each paper code
#                 ws.merge_cells(start_row=i, start_column=start_col, end_row=i, end_column=start_col + 1)
#                 ws.cell(row=i, column=start_col).value = f"Paper Code {paper_code}"
#                 ws.cell(row=i, column=start_col).font = Font(size=8)
#                 ws.cell(row=i, column=start_col).alignment = Alignment(horizontal='center')
#                 ws.cell(row=i, column=start_col + 2).value = count
#                 # Applying fill color to the cell
#                 ws.cell(row=i, column=start_col).fill = color_dict[paper_code]
                

#         # Save the workbook
#         wb.save(file_name)
#         messagebox.showinfo("Success", f"Exam desk cards saved successfully: {file_name}")

# # Create the Tkinter window
# root = tk.Tk()
# root.title("Generate Exam Desk Cards")

# # Function to generate exam desk cards
# def generate_cards():
#     generate_exam_desk_cards()

# # Create and place widgets
# generate_button = tk.Button(root, text="Generate Exam Desk Cards", command=generate_cards)
# generate_button.pack()

# root.mainloop()



import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog, messagebox

# Function to generate exam desk cards
def generate_exam_desk_cards(file_path):
    if not file_path:
        file_path = "test_seat_plan1.xlsx"  # Default file name
    
    # Read the data from the Excel file
    try:
        df = pd.read_excel(file_path)
    except Exception as e:
        messagebox.showerror("Error", f"Error reading file: {e}")
        return

    # Group students by exam venue and session
    grouped_data = df.groupby(['Venue', 'Session', 'Board'])

    # Define colors for each unique paper code
    color_dict = {}
    colors = [PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
              PatternFill(start_color='FFEB9B', end_color='FFEB9B', fill_type='solid'),
              PatternFill(start_color='C9FFC3', end_color='C9FFC3', fill_type='solid'),
              PatternFill(start_color='9BEBFF', end_color='9BEBFF', fill_type='solid'),
              PatternFill(start_color='FFC3E8', end_color='FFC3E8', fill_type='solid'),
              PatternFill(start_color='B8B8FF', end_color='B8B8FF', fill_type='solid'),
              PatternFill(start_color='D8BFD8', end_color='D8BFD8', fill_type='solid'),
              PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid'),
              PatternFill(start_color='F5DEB3', end_color='F5DEB3', fill_type='solid'),
              PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid'),
              PatternFill(start_color='808080', end_color='808080', fill_type='solid'),
              PatternFill(start_color='FFA07A', end_color='FFA07A', fill_type='solid'),
              PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid'),
              PatternFill(start_color='7B68EE', end_color='7B68EE', fill_type='solid'),
              PatternFill(start_color='FF6347', end_color='FF6347', fill_type='solid')]

    # Iterate over each group and generate exam desk cards
    for (venue, session, Board), group in grouped_data:
        # Create a file name for the exam desk cards
        file_name = f'Seat_Planing_for_{venue}_{session}.xlsx'
        
        # Create a workbook
        wb = Workbook()
        ws = wb.active
        
        # Add logo to the worksheet with adjusted position
        img = Image("British_Council_Logo.png")
        img.width = 150  # Adjust the width of the image (optional)
        img.height = 80  # Adjust the height of the image (optional)
        ws.add_image(img, 'A1')  # Set the anchor to 'A1' to position the image within the cell
        
        # Add headers and other details
        ws['A5'] = f"{Board} Examinations - {group['Date'].iloc[0].strftime('%Y-%m-%d')}"
        ws['A6'] = venue
        ws['A7'] = f"Session - {session}"
        ws['A8'] = f"Date: {group['Date'].iloc[0].strftime('%Y-%m-%d')}"
        
        paper_codes = ', '.join(map(str, group['Paper code'].unique()))
        ws['A9'] = f"Subject: {paper_codes}"

        # Add total candidate count
        ws.merge_cells('A10:B10')
        ws['A10'] = "Total candidate"
        ws['A10'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C10'] = len(group)

        # Initialize variables for row and column numbers
        start_row = 0
        start_col = 1
        
        # Iterate over candidate numbers and assign them to cells based on seat locator
        for _, row_data in group.iterrows():
            candidate_num = row_data['Candidate Number']
            seat_locator = row_data['Seat Locator']
            paper_code = row_data['Paper code']

            if pd.notnull(seat_locator):
                # Convert seat locator to row and column indices
                col = ord(seat_locator[0].lower()) - ord('a') + start_col  # Convert column letter to index
                row = int(seat_locator[1:]) + start_row  # Convert row number to index
                
                # Write candidate number to the specified cell
                ws.cell(row=row, column=col).value = candidate_num
                
                # Add border to the cell
                ws.cell(row=row, column=col).border = Border(left=Side(style='thin'), 
                                                              right=Side(style='thin'), 
                                                              top=Side(style='thin'), 
                                                              bottom=Side(style='thin'))
 
                # Assigning color to the cell based on the paper code
                if paper_code not in color_dict:
                    color_dict[paper_code] = colors.pop(0)
                
                # Applying fill color to the cell
                ws.cell(row=row, column=col).fill = color_dict[paper_code]

        # Add total candidate count for each paper code
        for i, paper_code in enumerate(group['Paper code'].unique(), start=11):
            count = len(group[group['Paper code'] == paper_code])
            if count > 0:
                # Write total count for each paper code
                ws.merge_cells(start_row=i, start_column=start_col, end_row=i, end_column=start_col + 1)
                ws.cell(row=i, column=start_col).value = f"Paper Code {paper_code}"
                ws.cell(row=i, column=start_col).font = Font(size=8)
                ws.cell(row=i, column=start_col).alignment = Alignment(horizontal='center')
                ws.cell(row=i, column=start_col + 2).value = count
                # Applying fill color to the cell
                ws.cell(row=i, column=start_col).fill = color_dict[paper_code]
                

        # Save the workbook
        wb.save(file_name)
        messagebox.showinfo("Success", f"Exam desk cards saved successfully: {file_name}")

# Create the Tkinter window
root = tk.Tk()
root.title("Generate Exam Desk Cards")

# Function to generate exam desk cards
def generate_cards():
    file_path = entry.get()
    generate_exam_desk_cards(file_path)

# Create and place widgets
frame = tk.Frame(root, padx=10, pady=10)
frame.pack()

label = tk.Label(frame, text="Generate Exam Desk Cards", font=("Arial", 16))
label.pack(pady=10)

entry_frame = tk.Frame(frame)
entry_frame.pack()

entry_label = tk.Label(entry_frame, text="Excel File Path:", font=("Arial", 12))
entry_label.pack(side=tk.LEFT)

entry = tk.Entry(entry_frame, font=("Arial", 12), width=30)
entry.pack(side=tk.LEFT, padx=(10, 0))

button = tk.Button(frame, text="Generate", font=("Arial", 12), command=generate_cards)
button.pack(pady=10)

root.mainloop()
