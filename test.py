# class ExamCenter:
#     def __init__(self, name):
#         self.name = name
#         self.rooms = []

#     def add_room(self, rows, columns):
#         room = Room(rows, columns)
#         self.rooms.append(room)

#     def display_seating_arrangement(self):
#         print(f"Seating arrangement for {self.name}:")
#         for i, room in enumerate(self.rooms, start=1):
#             print(f"Room {i}:")
#             room.display_seating_arrangement()


# class Room:
#     def __init__(self, rows, columns):
#         self.rows = rows
#         self.columns = columns
#         self.seats = [[None for _ in range(columns)] for _ in range(rows)]

#     def add_student(self, row, column, student_name):
#         if row < 1 or row > self.rows or column < 1 or column > self.columns:
#             print(f"Invalid seat: Row or column out of range for room {self.rows}x{self.columns}.")
#             return

#         if self.seats[row - 1][column - 1] is not None:
#             print(f"Seat ({row}, {column}) is already occupied.")
#             return

#         self.seats[row - 1][column - 1] = student_name
#         print(f"Assigned seat ({row}, {column}) to {student_name}.")

#     def display_seating_arrangement(self):
#         for row in range(self.rows):
#             for column in range(self.columns):
#                 seat = self.seats[row][column]
#                 if seat is None:
#                     print("|_|", end="\t")
#                 else:
#                     print(seat, end="\t")
#             print()


# # Example usage:
# exam_center = ExamCenter("Example Center")
# exam_center.add_room(rows=15, columns=5)
# exam_center.add_room(rows=12, columns=3)

# exam_center.rooms[0].add_student(row=5, column=3, student_name="Alice")
# exam_center.rooms[1].add_student(row=1, column=1, student_name="Bob")
# exam_center.rooms[1].add_student(row=2, column=2, student_name="Charlie")

# exam_center.display_seating_arrangement()



from openpyxl import Workbook

class ExamCenter:
    def __init__(self, name):
        self.name = name
        self.rooms = []

    def add_room(self, rows, columns):
        room = Room(rows, columns)
        self.rooms.append(room)

    def add_student_from_excel(self, excel_file):
        # Your existing code for adding students from Excel
        pass

    def save_seating_arrangement_to_excel(self, excel_file):
        wb = Workbook()
        for i, room in enumerate(self.rooms, start=1):
            ws = wb.create_sheet(title=f"Room {i}")
            for row in range(room.rows):
                for column in range(room.columns):
                    student_name = room.seats[row][column]
                    ws.cell(row=row + 1, column=column + 1, value=student_name)
        wb.save(excel_file)
        print(f"Seating arrangement saved to {excel_file}.")

class Room:
    def __init__(self, rows, columns):
        self.rows = rows
        self.columns = columns
        self.seats = [[None for _ in range(columns)] for _ in range(rows)]

    def add_student(self, row, column, student_name):
        if row < 1 or row > self.rows or column < 1 or column > self.columns:
            print(f"Invalid seat: Row or column out of range for room {self.rows}x{self.columns}.")
            return

        if self.seats[row - 1][column - 1] is not None:
            print(f"Seat ({row}, {column}) is already occupied.")
            return

        self.seats[row - 1][column - 1] = student_name
        print(f"Assigned seat ({row}, {column}) to {student_name}.")

    def display_seating_arrangement(self):
        for row in range(self.rows):
            for column in range(self.columns):
                seat = self.seats[row][column]
                if seat is None:
                    print("Empty", end="\t")
                else:
                    print(seat, end="\t")
            print()

# Example usage:
exam_center = ExamCenter("Example Center")
exam_center.add_room(rows=15, columns=5)

# Assuming you have added students from an Excel file using add_student_from_excel method
exam_center.add_student_from_excel("test_seat_plan1.xlsx")

# Saving the seating arrangement to an Excel file
exam_center.save_seating_arrangement_to_excel("seating_arrangement.xlsx")


# import pandas as pd
# from openpyxl.utils import column_index_from_string

# class ExamCenter:
#     def __init__(self, name):
#         self.name = name
#         self.rooms = []

#     def add_room(self, rows, columns):
#         room = Room(rows, columns)
#         self.rooms.append(room)

#     def add_student_from_excel(self, excel_file):
#         df = pd.read_excel(excel_file)
#         for index, row in df.iterrows():
#             candidate_number = row['Candidate Number']
#             seat_locator = row['Seat Locator']
#             if not pd.isnull(seat_locator):
#                 column_letter = seat_locator[0]  # Extracting the column letter (e.g., "B")
#                 row_index = int(seat_locator[1:])  # Extracting the row index (e.g., 15)
#                 column_index = column_index_from_string(column_letter)  # Converting column letter to index (e.g., 2)
#                 student_name = f"Student {candidate_number}"
#                 room_index = 1  # Assuming all students are in the first room
#                 self.rooms[room_index - 1].add_student(row_index, column_index, student_name)
#             else:
#                 print(f"No seat locator found for candidate {candidate_number}.")

#     def display_seating_arrangement(self):
#         print(f"Seating arrangement for {self.name}:")
#         for i, room in enumerate(self.rooms, start=1):
#             print(f"Room {i}:")
#             room.display_seating_arrangement()


# class Room:
#     def __init__(self, rows, columns):
#         self.rows = rows
#         self.columns = columns
#         self.seats = [[None for _ in range(columns)] for _ in range(rows)]

#     def add_student(self, row, column, student_name):
#         if row < 1 or row > self.rows or column < 1 or column > self.columns:
#             print(f"Invalid seat: Row or column out of range for room {self.rows}x{self.columns}.")
#             return

#         if self.seats[row - 1][column - 1] is not None:
#             print(f"Seat ({row}, {column}) is already occupied.")
#             return

#         self.seats[row - 1][column - 1] = student_name
#         print(f"Assigned seat ({row}, {column}) to {student_name}.")

#     def display_seating_arrangement(self):
#         for row in range(self.rows):
#             for column in range(self.columns):
#                 seat = self.seats[row][column]
#                 if seat is None:
#                     print("Empty", end="\t")
#                 else:
#                     print(seat, end="\t")
#             print()


# # Example usage:
# exam_center = ExamCenter("Example Center")
# exam_center.add_room(rows=15, columns=5)

# # Assuming candidate numbers and seat locators are listed in an Excel file named "test_seat_plan.xlsx"
# exam_center.add_student_from_excel("test_seat_plan1.xlsx")

# exam_center.display_seating_arrangement()







# import pandas as pd
# from openpyxl import Workbook

# class ExamCenter:
#     def __init__(self, name):
#         self.name = name
#         self.rooms = []

#     def add_room(self, rows, columns):
#         room = Room(rows, columns)
#         self.rooms.append(room)

#     def add_student_from_excel(self, excel_file):
#         df = pd.read_excel(excel_file)
#         print("Column Names")
#         for column in df.columns:
#             print(column)
#         room_index = 0
#         row_index = 1
#         column_index = 1
#         for index, row in df.iterrows():
#             candidate_number = row['Candidate Number']
#             student_name = f"Student {candidate_number}"
#             self.rooms[room_index].add_student(row_index, column_index, student_name)
#             # Move to the next seat
#             column_index += 1
#             if column_index > self.rooms[room_index].columns:
#                 column_index = 1
#                 row_index += 1
#                 if row_index > self.rooms[room_index].rows:
#                     row_index = 1
#                     room_index += 1
#                     if room_index >= len(self.rooms):
#                         print("Not enough rooms to assign all candidates.")
#                         break

#     def save_seating_arrangement_to_excel(self, excel_file):
#         wb = Workbook()
#         for i, room in enumerate(self.rooms, start=1):
#             ws = wb.create_sheet(title=f"Room {i}")
#             for row in range(room.rows):
#                 for column in range(room.columns):
#                     seat = room.seats[row][column]
#                     ws.cell(row=row+1, column=column+1, value=seat)
#         del wb['Sheet']  # Remove default sheet
#         wb.save(excel_file)


# class Room:
#     def __init__(self, rows, columns):
#         self.rows = rows
#         self.columns = columns
#         self.seats = [[None for _ in range(columns)] for _ in range(rows)]

#     def add_student(self, row, column, student_name):
#         if row < 1 or row > self.rows or column < 1 or column > self.columns:
#             print(f"Invalid seat: Row or column out of range for room {self.rows}x{self.columns}.")
#             return

#         if self.seats[row - 1][column - 1] is not None:
#             print(f"Seat ({row}, {column}) is already occupied.")
#             return

#         self.seats[row - 1][column - 1] = student_name

# # Example usage:
# exam_center = ExamCenter("Example Center")
# exam_center.add_room(rows=15, columns=5)
# exam_center.add_room(rows=12, columns=3)

# # Assuming candidate numbers are listed in an Excel file named "test_seat_plan.xlsx" with column "Candidate Number"
# exam_center.add_student_from_excel("test_seat_plan1.xlsx")

# # Save seating arrangement to an Excel file named "seating_arrangement.xlsx"
# exam_center.save_seating_arrangement_to_excel("seating_arrangement.xlsx")



from openpyxl import Workbook


wb = Workbook()
ws = wb.active