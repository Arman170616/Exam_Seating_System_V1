from ctypes import alignment
from django.shortcuts import render
from django.http import HttpResponseRedirect
from .models import Exam, Venue
import pandas as pd
from django.http import HttpResponse
from django.utils.datastructures import MultiValueDictKeyError


def import_data(request):
    if request.method == "POST" and 'Seat Plan Data_MJ2023.xlsx' in request.FILES:
        excel_file = request.FILES['Seat Plan Data_MJ2023.xlsx']
        
        if excel_file.name.endswith('.xlsx'):
            df = pd.read_excel(excel_file)
            for index, row in df.iterrows():
                Exam.objects.create(
                    Board=row['Board'],
                    Paper_code=row['Paper Code'],
                    Qualification=row['Qualification'],
                    Exam_type=row['Exam Type'],
                    Syllabus=row['Syllabus'],
                    Duration=row['Duration'],
                    Date=row['Date'],
                    Time_slot=row['Time Slot'],
                    Session=row['Session'],
                    Start_time=row['Start Time'],
                    End_time=row['End Time'],
                    Candidate_number=row['Candidate Number'],
                    unique_candidate=row['Unique Candidate']
                )
            return HttpResponseRedirect('/success/')
        else:
            return render(request, 'import.html', {'error': 'File format not supported'})
    
    return render(request, 'import.html')



def data_display(request):
    exams = Exam.objects.all()  
    imported_exams = Exam.objects.filter(unique_candidate__isnull=False)  # Filter imported data
    return render(request, 'data_display.html', {'exams': exams, 'imported_exams': imported_exams})



def upload_file(request):
    if request.method == 'POST':
        try:
            # Attempt to retrieve the uploaded file
            uploaded_file = request.FILES['file']
            
            # Check if the uploaded file is an Excel file
            if uploaded_file.name.endswith('.xlsx'):
                # Read the Excel file using pandas
                df = pd.read_excel(uploaded_file)
                
                # Convert DataFrame to HTML table
                html_table = df.to_html()
                
                # Pass the HTML table to the template for rendering
                return render(request, 'display_excel.html', {'html_table': html_table})
            else:
                # If the uploaded file is not an Excel file, display an error message
                return render(request, 'error.html', {'error_message': 'Invalid file format. Please upload an Excel file.'})
        except MultiValueDictKeyError:
            # If 'file' key is not found in request.FILES, display an error message
            return render(request, 'error.html', {'error_message': 'No file uploaded. Please choose a file to upload.'})
    else:
        # Render the upload file form
        return render(request, 'upload_file.html')




# FINAL OUTPUT





import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os


# Read the data from the Excel file
# df = pd.read_excel('TestData_BETA.xlsx')
try:
    df = pd.read_excel('27May.xlsx')
except FileNotFoundError:
    print("Failed to find '20May_M.xlsx'. Please check the file path.")
    raise

try:
    img = Image("British_Council_Logo.png")
except FileNotFoundError:
    print("Failed to find 'British_Council_Logo.png'. Please check the file path.") 
    raise

# Function to generate exam desk cards
def generate_exam_desk_cards(data):
    # Group students by exam venue, exam room, session, board, and date
    grouped_data = data.groupby(['Venue2', 'Exam Room', 'Session', 'Date'])

    # Define colors for each unique paper code
    colors = [PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
              PatternFill(start_color='FFEB9B', end_color='FFEB9B', fill_type='solid'),
              PatternFill(start_color='C9FFC3', end_color='C9FFC3', fill_type='solid'),
              PatternFill(start_color='9BEBFF', end_color='9BEBFF', fill_type='solid'),
              PatternFill(start_color='FFC3E8', end_color='FFC3E8', fill_type='solid'),
              PatternFill(start_color='B8B8FF', end_color='B8B8FF', fill_type='solid'),
              PatternFill(start_color='D8BFD8', end_color='D8BFD8', fill_type='solid'),
              PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid'),
              PatternFill(start_color='F5DEB3', end_color='F5DEB3', fill_type='solid'),
              PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid'),
              PatternFill(start_color='808080', end_color='808080', fill_type='solid'),
              PatternFill(start_color='FFA07A', end_color='FFA07A', fill_type='solid'),
              PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid'),
              PatternFill(start_color='7B68EE', end_color='7B68EE', fill_type='solid'),
              PatternFill(start_color='FF6347', end_color='FF6347', fill_type='solid')]

    # Iterate over each group and generate exam desk cards
    for ( venue, exam_room, session, date), group in grouped_data:
        date_str = date.strftime("%Y-%m-%d")
        directory_path = os.path.join('Exam_Seat_planner', venue, date_str)
        os.makedirs(directory_path, exist_ok=True)

        # Define file path
        file_name = f'Seat_Plan_{venue}_{exam_room}_{session}_{date_str}.xlsx'
        file_path = os.path.join(directory_path, file_name)


        # Create a workbook
        wb = Workbook()
        ws = wb.active

        img.width = 150  
        img.height = 80  
        ws.add_image(img, 'A1')  
        # Add headers and other details
        # ws['A5'] = f"Pearson Edexcel & Cambridge Int'l Examination - {date.strftime('%Y-%m-%d')}"
        ws['A5'] = f"Pearson Edexcel & Cambridge International Examination May/June 2024"
        ws['A6'] = f"Venue: {venue}, Exam Room: {exam_room}"
        ws['A7'] = f"Session - {session}"
        ws['A8'] = f"Date: {date.strftime('%Y-%m-%d')}"

        paper_codes = ', '.join(map(str, group['Paper code'].unique()))
        ws['A9'] = f"Subject: {paper_codes}"

        # Add total candidate count
        ws.merge_cells('A10:B10')
        ws['A10'] = "Total Candidate"
        ws['A10'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C10'] = len(group)

        # Initialize variables for row and column numbers
        start_row = 0
        start_col = 1

        # Initialize color index
        color_index = 0

        # Initialize color dictionary
        color_dict = {}

        # Iterate over candidate numbers and assign them to cells based on seat locator
        for _, row_data in group.iterrows():
            candidate_num = row_data['Candidate Number']
            seat_locator = row_data['Seat Locator']
            paper_code = row_data['Paper code']

            if pd.notnull(seat_locator):
                # Convert seat locator to row and column indices
                col = ord(seat_locator[0].lower()) - ord('a') + start_col  # Convert column letter to index
                row = int(seat_locator[1:]) + start_row  # Convert row number to index

                # Write candidate number to the specified cell
                ws.cell(row=row, column=col).value = candidate_num

                # Add border to the cell
                ws.cell(row=row, column=col).border = Border(left=Side(style='thin'),
                                                              right=Side(style='thin'),
                                                              top=Side(style='thin'),
                                                              bottom=Side(style='thin'))

                # Assigning color to the cell based on the paper code
                if paper_code not in color_dict:
                    color_dict[paper_code] = colors[color_index]
                    color_index = (color_index + 1) % len(colors)

                # Applying fill color to the cell
                ws.cell(row=row, column=col).fill = color_dict[paper_code]

        # Add total candidate count for each paper code
        for i, paper_code in enumerate(group['Paper code'].unique(), start=11):
            count = len(group[group['Paper code'] == paper_code])
            if count > 0:
                # Write total count for each paper code 
                ws.merge_cells(start_row=i, start_column=start_col, end_row=i, end_column=start_col + 1)
                ws.cell(row=i, column=start_col).value = f"Paper Code {paper_code}"
                ws.cell(row=i, column=start_col).font = Font(size=8)
                ws.cell(row=i, column=start_col).alignment = Alignment(horizontal='center')
                ws.cell(row=i, column=start_col + 2).value = count
                # Applying fill color to the cell
                ws.cell(row=i, column=start_col).fill = color_dict[paper_code]

        # Calculate the maximum number of columns needed based on the length of 'Seat Locator' values
        max_column = max([ord(seat_locator[0].lower()) - ord('a') + 1 for seat_locator in group['Seat Locator'] if pd.notnull(seat_locator)], default=0)

        # Add column numbers dynamically starting from cell A17
        for col_num in range(start_col, start_col + max_column):
            ws.cell(row=18, column=col_num).value = f"Column {col_num}"
            ws.cell(row=18, column=col_num).font = Font(size=9)

        # # Save the workbook
        wb.save(file_path)
        print(f"Successfully saved as: {file_path}")


# # Generate exam desk cards
generate_exam_desk_cards(df)




# import pandas as pd
# import os
# import logging
# import pdfkit
# from xlsxwriter import Workbook

# logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# try:
#     df = pd.read_excel('20May_M.xlsx')
# except FileNotFoundError:
#     logging.error("Failed to find '20May_M.xlsx'. Please check the file path.")
#     raise

# try:
#     logo_path = "British_Council_Logo.png"
#     if not os.path.exists(logo_path):
#         raise FileNotFoundError
# except FileNotFoundError:
#     logging.error("Failed to find 'British_Council_Logo.png'. Please check the file path.")
#     raise

# colors = ['FFC7CE', 'FFEB9B', 'C9FFC3', '9BEBFF', 'FFC3E8', 'B8B8FF', 'D8BFD8',
#           'C0C0C0', 'F5DEB3', 'FFD700', '808080', 'FFA07A', 'FFFACD', '7B68EE', 'FF6347']

# def generate_exam_desk_cards(data):
#     grouped_data = data.groupby(['Venue2', 'Exam Room', 'Session', 'Date'])

#     for (venue, exam_room, session, date), group in grouped_data:
#         try:
#             date_str = date.strftime("%Y-%m-%d")
#             directory_path = os.path.join('Exam_Seat_planner', venue, date_str)
#             os.makedirs(directory_path, exist_ok=True)

#             file_name = f'Seat_Plan_{venue}_{exam_room}_{session}_{date_str}.xlsx'
#             file_path = os.path.join(directory_path, file_name)

#             # Create a workbook and worksheet
#             wb = Workbook(file_path)
#             ws = wb.add_worksheet()

#             # Add headers and other details
#             ws.set_row(0, 60)
#             ws.insert_image('A1', logo_path, {'x_scale': 0.5, 'y_scale': 0.5})
#             ws.write('A5', "Pearson Edexcel & Cambridge International Examination May/June 2024")
#             ws.write('A6', f"Venue: {venue}, Exam Room: {exam_room}")
#             ws.write('A7', f"Session - {session}")
#             ws.write('A8', f"Date: {date_str}")

#             paper_codes = ', '.join(map(str, group['Paper code'].unique()))
#             ws.write('A9', f"Subject: {paper_codes}")

#             ws.merge_range('A10:B10', "Total Candidate")
#             ws.write('C10', len(group))

#             start_row = 0
#             start_col = 1

#             color_index = 0
#             color_dict = {}

#             for _, row_data in group.iterrows():
#                 candidate_num = row_data['Candidate Number']
#                 seat_locator = row_data['Seat Locator']
#                 paper_code = row_data['Paper code']

#                 if pd.notnull(seat_locator):
#                     col = ord(seat_locator[0].lower()) - ord('a') + start_col
#                     row = int(seat_locator[1:]) + start_row

#                     cell_format = wb.add_format({'border': 1})
#                     if paper_code not in color_dict:
#                         color_dict[paper_code] = colors[color_index]
#                         color_index = (color_index + 1) % len(colors)

#                     cell_format.set_bg_color(color_dict[paper_code])
#                     ws.write(row, col, candidate_num, cell_format)

#             wb.close()

#             html_file_path = file_path.replace('.xlsx', '.html')
#             pdf_file_path = file_path.replace('.xlsx', '.pdf')

#             df.to_html(html_file_path, index=False)

#             options = {
#                 'page-size': 'A4',
#                 'orientation': 'Portrait',
#                 'zoom': '0.85'
#             }

#             pdfkit.from_file(html_file_path, pdf_file_path, options=options)
#             os.remove(html_file_path)

#             logging.info(f"Successfully saved as: {pdf_file_path}")

#         except Exception as e:
#             logging.error(f"Failed to process group {(venue, exam_room, session, date)}: {e}")

# generate_exam_desk_cards(df)
